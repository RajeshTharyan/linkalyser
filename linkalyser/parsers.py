"""Extract searchable text from fetched document bytes."""

from io import BytesIO
from zipfile import BadZipFile

from bs4 import BeautifulSoup
from docx import Document
from docx.opc.exceptions import PackageNotFoundError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from PyPDF2 import PdfReader
from PyPDF2.errors import PdfReadError

# PDF pages are joined with form-feed so search_keywords can report page numbers.
PDF_PAGE_SEPARATOR = "\f"


def parse_pdf(content: bytes) -> str:
    try:
        reader = PdfReader(BytesIO(content))
        return PDF_PAGE_SEPARATOR.join(page.extract_text() or "" for page in reader.pages)
    except (PdfReadError, OSError, ValueError):
        return ""


def parse_word(content: bytes) -> str:
    try:
        document = Document(BytesIO(content))
        return "\n".join(paragraph.text for paragraph in document.paragraphs)
    except (PackageNotFoundError, OSError, ValueError, BadZipFile):
        return ""


def parse_excel(content: bytes) -> str:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True)
        lines = []
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows(values_only=True):
                lines.append(" ".join(str(cell) for cell in row if cell is not None))
        return "\n".join(lines)
    except (InvalidFileException, OSError, ValueError, BadZipFile):
        return ""


def parse_html(content: bytes | str) -> str:
    try:
        return BeautifulSoup(content, "html.parser").get_text() or ""
    except (ValueError, TypeError):
        return ""
